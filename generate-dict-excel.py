#!/usr/bin/env python3
"""
generate-dict-excel.py
----------------------
Generates the comprehensive MYOREHAB_Data_Dictionary.xlsx Excel workbook
containing data dictionary sheets for all entities in the MYOREHAB dataset.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def generate_myorehab_dict_excel(output_filename="MYOREHAB_Data_Dictionary.xlsx"):
    print("Building Data Dictionary structures...")

    # 0. Protocol Naming Syntax
    df_naming = pd.DataFrame([
        {"Element": "Folder Structure", "Value / Pattern": "emg-raw-renamed/", "Description & Categorization": "Directory containing raw/renamed sEMG and Kinematic CSV files", "Key Type": "-"},
        {"Element": "File Pattern", "Value / Pattern": "SNNN_PGA.csv", "Description & Categorization": "Standardized file naming syntax (e.g., S001_111.csv)", "Key Type": "-"},
        {"Element": "S", "Value / Pattern": "Subject Prefix", "Description & Categorization": "Literal letter 'S' indicating subject identifier", "Key Type": "-"},
        {"Element": "NNN", "Value / Pattern": "Subject Code", "Description & Categorization": "3-digit numeric ID (001-999) corresponding to subject_id (BR-CMP-SNNN)", "Key Type": "FK"},
        {"Element": "P", "Value / Pattern": "Arm Position", "Description & Categorization": "1 = Distal (AE: Above Elbow / Distal), 2 = Proximal (AF: Arm Flexed / Proximal)", "Key Type": "-"},
        {"Element": "G", "Value / Pattern": "Grasp Type", "Description & Categorization": "1 = One Finger Pinch (1FP), 2 = Two Finger Pinch (2FP), 3 = Full Grasp (FG)", "Key Type": "-"},
        {"Element": "A", "Value / Pattern": "Angle", "Description & Categorization": "1 = 0°, 2 = 45°, 3 = 90°, 4 = 135°, 5 = 180°", "Key Type": "-"}
    ])

    # 1. Centers Table
    df_centers = pd.DataFrame([
        {"Variable": "center_id", "Data Type": "VARCHAR(10)", "Unit/Format": "Alphanumeric", "Description": "Unique code for the research center (e.g., 'BR-CMP')", "Key Type": "PK"},
        {"Variable": "center_name", "Data Type": "VARCHAR(100)", "Unit/Format": "Text", "Description": "Full name of the institution/laboratory", "Key Type": "-"},
        {"Variable": "country", "Data Type": "VARCHAR(50)", "Unit/Format": "Text", "Description": "Country of the research center", "Key Type": "-"}
    ])

    # 2. Subjects Table
    df_subjects = pd.DataFrame([
        {"Variable": "subject_id", "Data Type": "VARCHAR(15)", "Unit/Format": "Alphanumeric", "Description": "Unique identifier (e.g., 'BR-CMP-S001')", "Key Type": "PK"},
        {"Variable": "center_id", "Data Type": "VARCHAR(10)", "Unit/Format": "Alphanumeric", "Description": "References the center", "Key Type": "FK"},
        {"Variable": "acronyms", "Data Type": "TEXT", "Unit/Format": "Text", "Description": "acronyms from name and family name", "Key Type": "-"},
        {"Variable": "age", "Data Type": "INTEGER", "Unit/Format": "Years", "Description": "Participant's age", "Key Type": "-"},
        {"Variable": "gender", "Data Type": "VARCHAR(20)", "Unit/Format": "Text", "Description": "Participant's gender", "Key Type": "-"},
        {"Variable": "weight", "Data Type": "NUMERIC(5,2)", "Unit/Format": "Kilograms (kg)", "Description": "Body weight", "Key Type": "-"},
        {"Variable": "height", "Data Type": "NUMERIC(5,2)", "Unit/Format": "Centimeters (cm)", "Description": "Height", "Key Type": "-"},
        {"Variable": "fa_circ", "Data Type": "NUMERIC(5,2)", "Unit/Format": "Centimeters (cm)", "Description": "Forearm circumference", "Key Type": "-"},
        {"Variable": "lateral_dominance", "Data Type": "VARCHAR(20)", "Unit/Format": "Text", "Description": "Handedness (Right/Left)", "Key Type": "-"},
        {"Variable": "laterality", "Data Type": "NUMERIC(5,2)", "Unit/Format": "Score", "Description": "Laterality index (e.g., Edinburgh Inventory)", "Key Type": "-"},
        {"Variable": "nhpeg_dominant", "Data Type": "NUMERIC(6,2)", "Unit/Format": "Seconds (s)", "Description": "NHPT time for dominant hand", "Key Type": "-"},
        {"Variable": "nhpeg_nondominant", "Data Type": "NUMERIC(6,2)", "Unit/Format": "Seconds (s)", "Description": "NHPT time for non-dominant hand", "Key Type": "-"},
        {"Variable": "injuries", "Data Type": "TEXT", "Unit/Format": "Text", "Description": "Clinical history or relevant upper-limb injuries", "Key Type": "-"},
        {"Variable": "injuries_description", "Data Type": "VARCHAR(20)", "Unit/Format": "Text", "Description": "type of injuries", "Key Type": "-"},
        {"Variable": "sport", "Data Type": "VARCHAR(20)", "Unit/Format": "Text", "Description": "Participant's sport", "Key Type": "-"},
        {"Variable": "sport_type", "Data Type": "VARCHAR(20)", "Unit/Format": "Text", "Description": "Type of sport practiced", "Key Type": "-"},
        {"Variable": "sport_frequency", "Data Type": "VARCHAR(20)", "Unit/Format": "Text", "Description": "Frequency of sport practiced in days in a week", "Key Type": "-"},
        {"Variable": "temporal_mark", "Data Type": "DATETIME", "Unit/Format": "YYYY-MM-DD HH:MM:SS", "Description": "Timestamp of form submission or data entry", "Key Type": "-"}
    ])

    # 3. Equipment Table
    df_equipment = pd.DataFrame([
        {"Variable": "equipment_id", "Data Type": "VARCHAR(15)", "Unit/Format": "Alphanumeric", "Description": "Unique code for the hardware setup", "Key Type": "PK"},
        {"Variable": "center_id", "Data Type": "VARCHAR(10)", "Unit/Format": "Alphanumeric", "Description": "References the research center", "Key Type": "FK"},
        {"Variable": "signal_type", "Data Type": "VARCHAR(10)", "Unit/Format": "Text", "Description": "Signal modality ('EMG' or 'KINEMATICS')", "Key Type": "-"},
        {"Variable": "brand", "Data Type": "VARCHAR(50)", "Unit/Format": "Text", "Description": "Manufacturer (e.g., 'In-house UNICAMP')", "Key Type": "-"},
        {"Variable": "model", "Data Type": "VARCHAR(50)", "Unit/Format": "Text", "Description": "Hardware model", "Key Type": "-"},
        {"Variable": "channel_layout", "Data Type": "VARCHAR(50)", "Unit/Format": "Text", "Description": "Physical distribution (e.g., '2x64 matrices', '4x32 bands')", "Key Type": "-"},
        {"Variable": "total_channels", "Data Type": "INTEGER", "Unit/Format": "Count", "Description": "Total data channels", "Key Type": "-"}
    ])

    # 4. Tasks Table
    df_tasks = pd.DataFrame([
        {"Variable": "task_id", "Data Type": "VARCHAR(20)", "Unit/Format": "Alphanumeric", "Description": "Unique code combining subject and condition (e.g., 'BR-CMP-S001_M111')", "Key Type": "PK"},
        {"Variable": "subject_id", "Data Type": "VARCHAR(15)", "Unit/Format": "Alphanumeric", "Description": "References the participant", "Key Type": "FK"},
        {"Variable": "condition_code", "Data Type": "VARCHAR(5)", "Unit/Format": "Categorical", "Description": "Protocol condition (M111 to M235) encoding PGA: P=Position, G=Grasp, A=Angle", "Key Type": "-"},
        {"Variable": "duration", "Data Type": "NUMERIC(5,2)", "Unit/Format": "Seconds (s)", "Description": "Standardized duration of the trial (e.g., 30.0 s)", "Key Type": "-"}
    ])

    # 5. Recordings Table
    df_recordings = pd.DataFrame([
        {"Variable": "recording_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "Unique identifier for the recording file", "Key Type": "PK"},
        {"Variable": "task_id", "Data Type": "VARCHAR(20)", "Unit/Format": "Alphanumeric", "Description": "References the mirrored task/condition", "Key Type": "FK"},
        {"Variable": "equipment_id", "Data Type": "VARCHAR(15)", "Unit/Format": "Alphanumeric", "Description": "References the hardware used", "Key Type": "FK"},
        {"Variable": "sampling_freq", "Data Type": "NUMERIC(8,2)", "Unit/Format": "Hertz (Hz)", "Description": "Actual sampling frequency", "Key Type": "-"}
    ])

    # 6. EMG Signals Table (128 Channels)
    emg_rows = [
        {"Variable": "emg_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "Unique identifier for the EMG signal metadata entry", "Key Type": "PK"},
        {"Variable": "recording_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "References the recording session", "Key Type": "FK"},
        {"Variable": "equipment_id", "Data Type": "VARCHAR(15)", "Unit/Format": "Alphanumeric", "Description": "References the EMG acquisition equipment", "Key Type": "FK"},
        {"Variable": "file_path", "Data Type": "VARCHAR(255)", "Unit/Format": "Relative Path", "Description": "Relative file path to the raw/renamed sEMG CSV file", "Key Type": "-"}
    ]
    for i in range(1, 65):
        emg_rows.append({
            "Variable": f"flex_ch{i}", "Data Type": "FLOAT64", "Unit/Format": "uV / Digital",
            "Description": f"HD-sEMG channel {i} placed over forearm flexor muscles", "Key Type": "-"
        })
    for i in range(1, 65):
        emg_rows.append({
            "Variable": f"exte_ch{i}", "Data Type": "FLOAT64", "Unit/Format": "uV / Digital",
            "Description": f"HD-sEMG channel {i} placed over forearm extensor muscles", "Key Type": "-"
        })
    df_emg = pd.DataFrame(emg_rows)

    # 7. Kinematics 3D Table
    kin_rows = [
        {"Variable": "kin_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "Unique identifier for the kinematic 3D metadata entry", "Key Type": "PK"},
        {"Variable": "recording_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "References the recording session", "Key Type": "FK"},
        {"Variable": "equipment_id", "Data Type": "VARCHAR(15)", "Unit/Format": "Alphanumeric", "Description": "References the kinematic acquisition equipment", "Key Type": "FK"},
        {"Variable": "file_path", "Data Type": "VARCHAR(255)", "Unit/Format": "Relative Path", "Description": "Relative file path to the 3D kinematics CSV file", "Key Type": "-"},
        {"Variable": "fnum", "Data Type": "INT64", "Unit/Format": "Frame Index", "Description": "Sequential temporal frame index", "Key Type": "-"}
    ]

    joints = [
        ("wrist", "Wrist"),
        ("thumb_cmc", "Thumb Cmc"), ("thumb_mcp", "Thumb Mcp"), ("thumb_ip", "Thumb Ip"), ("thumb_tip", "Thumb Tip"),
        ("index_finger_mcp", "Index Finger Mcp"), ("index_finger_pip", "Index Finger Pip"), ("index_finger_dip", "Index Finger Dip"), ("index_finger_tip", "Index Finger Tip"),
        ("middle_finger_mcp", "Middle Finger Mcp"), ("middle_finger_pip", "Middle Finger Pip"), ("middle_finger_dip", "Middle Finger Dip"), ("middle_finger_tip", "Middle Finger Tip"),
        ("ring_finger_mcp", "Ring Finger Mcp"), ("ring_finger_pip", "Ring Finger Pip"), ("ring_finger_dip", "Ring Finger Dip"), ("ring_finger_tip", "Ring Finger Tip"),
        ("pinky_mcp", "Pinky Mcp"), ("pinky_pip", "Pinky Pip"), ("pinky_dip", "Pinky Dip"), ("pinky_tip", "Pinky Tip")
    ]

    for prefix, label in joints:
        kin_rows.append({"Variable": f"{prefix}_x", "Data Type": "FLOAT64", "Unit/Format": "mm / px", "Description": f"3D spatial X coordinate for {label}", "Key Type": "-"})
        kin_rows.append({"Variable": f"{prefix}_y", "Data Type": "FLOAT64", "Unit/Format": "mm / px", "Description": f"3D spatial Y coordinate for {label}", "Key Type": "-"})
        kin_rows.append({"Variable": f"{prefix}_z", "Data Type": "FLOAT64", "Unit/Format": "mm / px", "Description": f"3D spatial Z coordinate for {label}", "Key Type": "-"})
        kin_rows.append({"Variable": f"{prefix}_error", "Data Type": "FLOAT64", "Unit/Format": "Pixels (px)", "Description": f"Reprojection error in 2D pixels for {label}", "Key Type": "-"})
        kin_rows.append({"Variable": f"{prefix}_ncams", "Data Type": "INT64", "Unit/Format": "Count (0-5)", "Description": f"Number of cameras contributing to triangulation for {label}", "Key Type": "-"})
        kin_rows.append({"Variable": f"{prefix}_score", "Data Type": "FLOAT64", "Unit/Format": "Score (0-1)", "Description": f"Minimum 2D detection confidence score for {label}", "Key Type": "-"})

    for i in range(3):
        for j in range(3):
            kin_rows.append({"Variable": f"m_{i}{j}", "Data Type": "FLOAT64", "Unit/Format": "Matrix Element", "Description": f"Element ({i},{j}) of the 3x3 coordinate rotation matrix M", "Key Type": "-"})

    for i in range(3):
        kin_rows.append({"Variable": f"center_{i}", "Data Type": "FLOAT64", "Unit/Format": "Spatial Coords", "Description": f"Coordinate {i} of the 3D reference system origin center", "Key Type": "-"})

    df_kinematics = pd.DataFrame(kin_rows)

    # 8. Videos 3D Table
    df_videos = pd.DataFrame([
        {"Variable": "video_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "Unique identifier for the 3D reconstructed video metadata entry", "Key Type": "PK"},
        {"Variable": "recording_id", "Data Type": "UUID", "Unit/Format": "UUID v4", "Description": "References the recording session", "Key Type": "FK"},
        {"Variable": "fps", "Data Type": "INTEGER", "Unit/Format": "Frames per second", "Description": "Frame rate of the reconstructed MP4 video", "Key Type": "-"},
        {"Variable": "file_path", "Data Type": "VARCHAR(255)", "Unit/Format": "Relative Path", "Description": "Relative file path to the 3D reconstructed MP4 video file", "Key Type": "-"}
    ])
    sheets = {
        "0_Protocol_Naming": df_naming,
        "1_Centers": df_centers,
        "2_Subjects": df_subjects,
        "3_Equipment": df_equipment,
        "4_Tasks": df_tasks,
        "5_Recordings": df_recordings,
        "6_EMG_Signals": df_emg,
        "7_Kinematics_3D": df_kinematics,
        "8_Videos_3D": df_videos
    }

    # OpenPyXL Styling setup
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    border_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

    print("Generating formatted Excel sheets...")
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            for c_idx, _ in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = header_font if r_idx == 1 else data_font
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.border = cell_border
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill
                    if df.columns[c_idx - 1] in ["Key Type", "Data Type", "Unit/Format"]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_filename)
    print(f" Successfully generated '{output_filename}' with {len(sheets)} sheets!")

if __name__ == "__main__":
    generate_myorehab_dict_excel()